from openpyxl import *
import csv
import xlrd
import re

wb = xlrd.open_workbook('Arts.xls')
sh = wb.sheet_by_index(0)

a = sh.cell_value(17,1)

print(a)

#for col in range(0,sh.ncols):
 #   print(sh.cell_value(1,col))

data = [[sh.cell_value(r,c) for c in range (sh.ncols)]
for r in range (sh.nrows)]



list = []
for i in range(0,7):
    type = 'arts'
    year = data[3][2+i*4]
    sth = sh.col_values(0,6,7) + sh.col_values(0,9,13)
    prg = sh.col_values(2+i*4,6,7) + sh.col_values(2+i*4,9,13)
    rg = sh.col_values(3+i*4,6,7) + sh.col_values(3+i*4,9,13)
    resp = sh.col_values(4+i*4,6,7) + sh.col_values(4+i*4,9,13)



    empty = "\\N"
    lempty ="\\N"
    for i in range(0,5):
        prog = re.compile("20\d\d/20\d\d")
        search = re.search(prog, str(year))
        if search:
            reg_year = search.group()
        else:
            reg_year = year[0:5] + "20" + year[5:7]
        sublist=[type,reg_year,str(sth[i]).replace(',',''),str(prg[i]).replace(',',''),
                str(rg[i]).replace(',',''),str(resp[i]).replace(',','')]
        list.append(sublist)

for i in range(0,3):
    type = 'arts'
    year = data[3][30+i*14]
    sth = sh.col_values(0,6,7) + sh.col_values(0,9,13)
    prg = sh.col_values(30+i*14,6,7) + sh.col_values(30+i*14,9,13)
    rg = sh.col_values(32+i*14,6,7) + sh.col_values(32+i*14,9,13)
    resp = sh.col_values(42+i*14,6,7) + sh.col_values(42+i*14,9,13)



    empty = "\\N"
    lempty ="\\N"
    for i in range(0,5):
        prog = re.compile("20\d\d/20\d\d")
        search = re.search(prog, str(year))
        if search:
            reg_year = search.group()
        else:
            reg_year = year[0:5] + "20" + year[5:7]
        sublist=[type,reg_year,str(sth[i]).replace(',',''),str(prg[i]).replace(',',''),
                str(rg[i]).replace(',',''),str(resp[i]).replace(',','')]
        list.append(sublist)



print(list.__sizeof__())

with open('tp_arts_freq.csv', 'w', newline='') as f:
    c = csv.writer(f)
    for i in list:
        c.writerow(i)



'''
wb =Workbook()
wb = load_workbook('Arts.xls')
#sh = wb.get_index(1)
#wb.save('lala.csv')
sh = wb.get_active_sheet()

#tuple(sh.iter_rows('A1:C2'))
names =[]


for row in sh.iter_rows('A1:W686'):
    for cell in row:
        names.append(cell.value)




names2 = names[23::23]
alt_name = names[24::23]
area = names[26::23]
region = names[27::23]
cat = names[28::23]
subcat = []
country = ['England']
local = names[44::23]
website = names[45::23]


year = '2017/2018'
funds = names[36::23]
prog = names[25::23]
source = names[38::23]
notes = website = names[43::23]
names2 = names[23::23]


list = []
empty = "\\N"
lempty ="\\N"
for i in range(0,685):
    sublist=[year,str(prog[i]).replace(',',''),str(source[i]).replace(',',''), str(funds[i]).replace(',',''),str(notes[i]).replace(',',''),
             str(names2[i]).replace(',','')]
    list.append(sublist)

print(list.__sizeof__())

with open('tp_arts.csv', 'w') as f:
    c = csv.writer(f)
    for i in list:
        c.writerow(i)



#print(lala)

 with open('test.csv', 'w') as f:
    c = csv.writer(f)
    for r in sh.rows:
        #.encode('utf-8')
        c.writerow([cell.value for cell in r])

                '''