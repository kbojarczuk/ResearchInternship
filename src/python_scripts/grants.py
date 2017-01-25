from openpyxl import *
import csv
import xlrd
import re


#list of ids

wb2 =Workbook()
wb2 = load_workbook('list.xlsx')
#sh = wb.get_index(1)
sh = wb2.get_active_sheet()

#tuple(sh.iter_rows('A1:C2'))
names =[]


for row in sh.iter_rows('A1:W686'):
    for cell in row:
        names.append(cell.value)

id=[]
for i in range(1,700):
    id.append(i)

names2 = names[23::23]

ids = dict(zip(names2,id))


#grants ---------------------------------------

wb = xlrd.open_workbook('grants_14.xls')
sh = wb.sheet_by_index(0)


# [row][col]
list = []
year = [sh.cell_value(r,0) for r in range (7,sh.nrows)]
name = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
name2 = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
source = [sh.cell_value(r,2) for r in range (7,sh.nrows)]
prog = [sh.cell_value(r,3) for r in range (7,sh.nrows)]
heading = [sh.cell_value(r,4) for r in range (7,sh.nrows)]
grant = [sh.cell_value(r,5) for r in range (7,sh.nrows)]
desc = [sh.cell_value(r,6) for r in range (7,sh.nrows)]
artform = [sh.cell_value(r,7) for r in range (7,sh.nrows)]
local = [sh.cell_value(r,8) for r in range (7,sh.nrows)]
reg = [sh.cell_value(r,9) for r in range (7,sh.nrows)]
cons = [sh.cell_value(r,10) for r in range (7,sh.nrows)]
ward = [sh.cell_value(r,11) for r in range (7,sh.nrows)]

ids_inst=[]
for n in name:
    if ids.get(n):
        ids_inst.append(ids.get(n))
    else:
        ids_inst.append("\\N")


empty = "\\N"
lempty ="\\N"
for i in range(0,5658):

    a =  desc[i].replace(',','')
    b = a.replace('\n','')
    c = b.replace('\r','')
    descf = c.replace('\"','')

    sublist = [empty,str(year[i]),name2[i].replace(',',''),source[i].replace(',',''),prog[i].replace(',',''),heading[i].replace(',',''),grant[i],descf,artform[i].replace(',',''),
               local[i].replace(',',''),reg[i].replace(',',''),cons[i].replace(',',''),ward[i].replace(',','')]
    list.append(sublist)
        #sublist=[reg_year,str(sth[i]).replace(',',''),str(prg[i]).replace(',',''),
         # str(rg[i]).replace(',',''),str(resp[i]).replace(',','')]


wb = xlrd.open_workbook('grants_15_1.xls')
sh = wb.sheet_by_index(0)


# [row][col]
year = [sh.cell_value(r,0) for r in range (7,sh.nrows)]
name = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
name2 = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
source = [sh.cell_value(r,2) for r in range (7,sh.nrows)]
prog = [sh.cell_value(r,3) for r in range (7,sh.nrows)]
heading = [sh.cell_value(r,4) for r in range (7,sh.nrows)]
grant = [sh.cell_value(r,5) for r in range (7,sh.nrows)]
desc = [sh.cell_value(r,6) for r in range (7,sh.nrows)]
artform = [sh.cell_value(r,7) for r in range (7,sh.nrows)]
local = [sh.cell_value(r,8) for r in range (7,sh.nrows)]
reg = [sh.cell_value(r,9) for r in range (7,sh.nrows)]
cons = [sh.cell_value(r,10) for r in range (7,sh.nrows)]
ward = [sh.cell_value(r,11) for r in range (7,sh.nrows)]

ids_inst=[]
for n in name:
    if ids.get(n):
        ids_inst.append(ids.get(n))
    else:
        ids_inst.append("\\N")


empty = "\\N"
lempty ="\\N"
for i in range(0,2194):

    a =  desc[i].replace(',','')
    b = a.replace('\n','')
    c = b.replace('\r','')
    descf = c.replace('\"','')

    sublist = [empty,str(year[i]),name2[i].replace(',',''),source[i].replace(',',''),prog[i].replace(',',''),heading[i].replace(',',''),grant[i],descf,artform[i].replace(',',''),
               local[i].replace(',',''),reg[i].replace(',',''),cons[i].replace(',',''),ward[i].replace(',','')]
    list.append(sublist)

wb = xlrd.open_workbook('grants_15_2.xls')
sh = wb.sheet_by_index(0)


# [row][col]
year = [sh.cell_value(r,0) for r in range (7,sh.nrows)]
name = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
name2 = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
source = [sh.cell_value(r,2) for r in range (7,sh.nrows)]
prog = [sh.cell_value(r,3) for r in range (7,sh.nrows)]
heading = [sh.cell_value(r,4) for r in range (7,sh.nrows)]
grant = [sh.cell_value(r,5) for r in range (7,sh.nrows)]
desc = [sh.cell_value(r,6) for r in range (7,sh.nrows)]
artform = [sh.cell_value(r,7) for r in range (7,sh.nrows)]
local = [sh.cell_value(r,8) for r in range (7,sh.nrows)]
reg = [sh.cell_value(r,9) for r in range (7,sh.nrows)]
cons = [sh.cell_value(r,10) for r in range (7,sh.nrows)]
ward = [sh.cell_value(r,11) for r in range (7,sh.nrows)]

ids_inst=[]
for n in name:
    if ids.get(n):
        ids_inst.append(ids.get(n))
    else:
        ids_inst.append("\\N")


empty = "\\N"
lempty ="\\N"
for i in range(0,3339):

    a =  desc[i].replace(',','')
    b = a.replace('\n','')
    c = b.replace('\r','')
    descf = c.replace('\"','')

    sublist = [empty,str(year[i]),name2[i].replace(',',''),source[i].replace(',',''),prog[i].replace(',',''),heading[i].replace(',',''),grant[i],descf,artform[i].replace(',',''),
               local[i].replace(',',''),reg[i].replace(',',''),cons[i].replace(',',''),ward[i].replace(',','')]
    list.append(sublist)

wb = xlrd.open_workbook('grants_15_3.xls')
sh = wb.sheet_by_index(0)


# [row][col]
year = [sh.cell_value(r,0) for r in range (7,sh.nrows)]
name = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
name2 = [sh.cell_value(r,1) for r in range (7,sh.nrows)]
source = [sh.cell_value(r,2) for r in range (7,sh.nrows)]
prog = [sh.cell_value(r,3) for r in range (7,sh.nrows)]
heading = [sh.cell_value(r,4) for r in range (7,sh.nrows)]
grant = [sh.cell_value(r,5) for r in range (7,sh.nrows)]
desc = [sh.cell_value(r,6) for r in range (7,sh.nrows)]
artform = [sh.cell_value(r,7) for r in range (7,sh.nrows)]
local = [sh.cell_value(r,8) for r in range (7,sh.nrows)]
reg = [sh.cell_value(r,9) for r in range (7,sh.nrows)]
cons = [sh.cell_value(r,10) for r in range (7,sh.nrows)]
ward = [sh.cell_value(r,11) for r in range (7,sh.nrows)]

ids_inst=[]
for n in name:
    if ids.get(n):
        ids_inst.append(ids.get(n))
    else:
        ids_inst.append("\\N")


empty = "\\N"
lempty ="\\N"
for i in range(0,4330):

    a =  desc[i].replace(',','')
    b = a.replace('\n','')
    c = b.replace('\r','')
    descf = c.replace('\"','')

    sublist = [empty,str(year[i]),name2[i].replace(',',''),source[i].replace(',',''),prog[i].replace(',',''),heading[i].replace(',',''),grant[i],descf,artform[i].replace(',',''),
               local[i].replace(',',''),reg[i].replace(',',''),cons[i].replace(',',''),ward[i].replace(',','')]
    list.append(sublist)

print(list.__sizeof__())

with open('grants.csv', 'w', newline='') as f:
    c = csv.writer(f)
    for i in list:
        c.writerow(i)


