from openpyxl import *
import csv
import uuid
import itertools
import xlrd


#getting local authorities ids 
wb2 = xlrd.open_workbook('local2012.xls')
sh2 = wb2.sheet_by_index(2)


# [row][col]
list = []
ecode = [sh2.cell_value(r,0) for r in range(9,453)]
name_local = [sh2.cell_value(r,1) for r in range(9,453)]
print(name_local)


#opening list of institutions
wb =Workbook()
wb = load_workbook('list.xlsx')
#sh = wb.get_index(1)
wb.save('lala.csv')
sh = wb.get_active_sheet()

#tuple(sh.iter_rows('A1:C2'))
names =[]

#getting all cells
for row in sh.iter_rows('A1:W686'):
    for cell in row:
        names.append(cell.value)

id=[]
for i in range(1,700):
    id.append(i)

names2 = names[23::23]
alt_name = names[24::23]
area = names[26::23]
region = names[27::23]
cat = names[28::23]
subcat = []
country = ['England']
local = names[44::23]
website = names[45::23]
lat = []
lng = []
local_ids=[]
#putting institution ids in a dictionary
ids = dict(zip(names2,id))
print(ids)

list = []
empty = "\\N"
lempty ="\\N"
for i in range(0,685):
    lala = False
    for j in range(0,443):
        if local[i] in name_local[j]:
            local_ids.append(ecode[j])
            lala = True
            break
        elif local[i]=='Powys':
            local_ids.append('02')
            lala = True
            break
        elif local[i] == 'Cardiff':
            local_ids.append('01')
            lala = True
            break
    if lala == False:
        local_ids.append('no')


    sublist=[id[i],str(names2[i]).replace(',',''),str(alt_name[i]).replace(',',''), str(area[i]).replace(',',''),str(region[i]).replace(',',''),str(cat[i]).replace(',',''),empty,'England',
             str(local[i]).replace(',',''),str(website[i]).replace(',',''),empty,empty,local_ids[i][1:5]]
    list.append(sublist)

print(list.__sizeof__())

with open('ac_inst.csv', 'w', newline='') as f:
    c = csv.writer(f)
    for i in list:
        c.writerow(i)