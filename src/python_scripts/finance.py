from openpyxl import *
import csv

#list of ids

wb2 =Workbook()
wb2 = load_workbook('list.xlsx')
#sh = wb.get_index(1)
sh = wb2.get_active_sheet()

#tuple(sh.iter_rows('A1:C2'))
names =[]


for row in sh.iter_rows('A1:W686'):
    for cell in row:
        names.append(cell.value)

id=[]
for i in range(1,700):
    id.append(i)

names2 = names[23::23]

ids = dict(zip(names2,id))
print(ids)

#finance

wb =Workbook()
wb = load_workbook('list.xlsx')
#sh = wb.get_index(1)
#wb.save('lala.csv')
sh = wb.get_active_sheet()

#tuple(sh.iter_rows('A1:C2'))
names =[]


for row in sh.iter_rows('A1:W686'):
    for cell in row:
        names.append(cell.value)



year = '2012/2013'
funds = names[30::23]
prog = names[25::23]
source = names[38::23]
notes2 = website = names[43::23]
notes = [x if x is not None else "NULL" for x in notes2]
names2 = names[23::23]
ids_order =[]

for name in names2:
    ids_order.append(ids.get(name))


list = []
empty = "\\N"
lempty ="\\N"
for i in range(0,685):
    sublist=[year,str(prog[i]).replace(',',''),str(source[i]).replace(',',''), str(funds[i]).replace(',',''),str(notes[i]).replace(',',''),
             ids_order[i]]
    list.append(sublist)

print(list.__sizeof__())

with open('funds1213.csv', 'w', newline='') as f:
    c = csv.writer(f)
    for i in list:
        c.writerow(i)



#print(lala)

''' with open('test.csv', 'w') as f:
    c = csv.writer(f)
    for r in sh.rows:
        #.encode('utf-8')
        c.writerow([cell.value for cell in r]) '''